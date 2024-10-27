from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import requests
from bs4 import BeautifulSoup
import re
from datetime import datetime
import logging
from pymongo import MongoClient, UpdateOne
from bson import json_util
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO

# Inicialización de la aplicación Flask y configuración de CORS
app = Flask(__name__)
CORS(app)  # Permite peticiones desde cualquier origen

# Configuración del sistema de logs
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Conexión a MongoDB
client = MongoClient('mongodb://localhost:27017/')
db = client['mercadolibre_db']
collection = db['productos']

def extraer_detalles_producto(url_producto, headers):
    """
    Extrae los detalles de un producto de MercadoLibre dado su URL
    """
    try:
        r_producto = None
        # Intentar obtener la página hasta 3 veces
        for _ in range(3):
            r_producto = requests.get(url_producto, headers=headers)
            if r_producto.status_code == 200:
                break

        if r_producto and r_producto.status_code == 200:
            soup_producto = BeautifulSoup(r_producto.content, 'html.parser')

            # Extraer información del vendedor
            vendedor_tag = soup_producto.find('div', class_='ui-pdp-seller__header__title')
            vendedor = vendedor_tag.get_text(strip=True) if vendedor_tag else 'N/A'
            
            # Inicializar variables de precios
            precio_original = 'N/A'
            precio_con_descuento = 'N/A'
            descuento = 'N/A'

            # Extraer precio original
            precio_original_tag = soup_producto.find('s', class_='andes-money-amount ui-pdp-price__part ui-pdp-price__original-value andes-money-amount--previous andes-money-amount--cents-superscript andes-money-amount--compact')
            if precio_original_tag:
                precio_original = precio_original_tag.text.strip()

            # Extraer precio con descuento
            precio_con_descuento_tag = soup_producto.find('div', class_='ui-pdp-price__second-line')
            if precio_con_descuento_tag:
                precio_descuento = precio_con_descuento_tag.find('span', class_='andes-money-amount__fraction')
                if precio_descuento:
                    precio_con_descuento = precio_descuento.text.strip()

            # Extraer porcentaje de descuento
            descuento_tag = soup_producto.find('span', class_='ui-pdp-price__second-line__label')
            if descuento_tag:
                descuento = descuento_tag.text.strip()

            # Extraer información de cuotas
            cuotas_pago = 'N/A'
            cuotas_container = soup_producto.find('div', class_='ui-pdp-payment')
            if cuotas_container:
                cuotas_text = cuotas_container.get_text(strip=True)
                cuotas_match = re.search(r'(\d+)x\s*(\$[\d,.]+)\s*(sin interés|con interés)?', cuotas_text)
                if cuotas_match:
                    cuotas_pago = f"{cuotas_match.group(1)}x {cuotas_match.group(2)} {cuotas_match.group(3) or ''}"

            # Extraer información de meses sin intereses
            meses_sin_intereses = 'N/A'
            meses_sin_intereses_tag = soup_producto.find('span', class_='ui-pdp-color--GREEN')
            if meses_sin_intereses_tag:
                meses_sin_intereses = meses_sin_intereses_tag.get_text(strip=True)

            # Extraer información de envío
            envio_tag = soup_producto.find('p', class_='ui-pdp-color--BLACK ui-pdp-family--REGULAR ui-pdp-media__title')
            envio = envio_tag.get_text(strip=True) if envio_tag else 'N/A'

            # Extraer cantidad vendida
            cantidad_vendida = 'N/A'
            cantidad_tag = soup_producto.find('span', class_='ui-pdp-subtitle')
            if cantidad_tag:
                cantidad_match = re.search(r'(\d+)\s+vendidos?', cantidad_tag.text)
                if cantidad_match:
                    cantidad_vendida = cantidad_match.group(1)

            # Extraer URL de la imagen
            imagen_tag = soup_producto.find('img', class_='ui-pdp-image ui-pdp-gallery__figure__image')
            imagen = imagen_tag['src'] if imagen_tag and 'src' in imagen_tag.attrs else 'N/A'

            # Devolver todos los datos extraídos
            return {
                'vendedor': vendedor,
                'precio_original': precio_original,
                'precio_con_descuento': precio_con_descuento,
                'descuento': descuento,
                'cuotas': cuotas_pago,
                'meses_sin_intereses': meses_sin_intereses,
                'envios': envio,
                'cantidad_vendida': cantidad_vendida,
                'imagenes': imagen
            }
        return None
    except Exception as e:
        logger.error(f"Error al procesar la página del producto {url_producto}: {e}")
        return None

def buscar_producto_api(producto):
    """
    Busca productos en la API de MercadoLibre
    """
    try:
        url = f'https://api.mercadolibre.com/sites/MLM/search?q={producto}'
        response = requests.get(url)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        logger.error(f"Error al buscar producto en la API: {e}")
        return None

def guardar_productos_en_db(datos_productos):
    """
    Guarda o actualiza los productos en MongoDB
    """
    if datos_productos:
        try:
            operations = []
            for producto in datos_productos:
                filter_query = {"url_producto": producto["url_producto"]}
                update_query = {"$set": producto}
                operations.append(UpdateOne(filter_query, update_query, upsert=True))
            
            result = collection.bulk_write(operations)
            logger.info(f"Se han guardado/actualizado {result.upserted_count + result.modified_count} productos en la base de datos.")
        except Exception as e:
            logger.error(f"Error al insertar/actualizar productos en MongoDB: {e}")

@app.route('/mercadolibre', methods=['POST', 'OPTIONS'])
def buscar_productos():
    """
    Endpoint principal para buscar productos
    """
    if request.method == 'OPTIONS':
        return '', 200
        
    try:
        data = request.get_json()
        if not data or 'producto' not in data:
            return jsonify({'error': 'No se proporcionó un producto'}), 400
        
        producto = data['producto']
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }

        start_time = datetime.now()
        productos_a_guardar = []

        resultado_api = buscar_producto_api(producto)
        if not resultado_api:
            return jsonify({'error': 'Error al buscar productos'}), 500

        for item in resultado_api['results']:
            url_producto = item['permalink']
            detalles = extraer_detalles_producto(url_producto, headers)

            if detalles:
                producto_data = {
                    "titulo": item['title'],
                    "url_producto": url_producto,
                    **detalles,
                    "fecha_extraccion": datetime.now()
                }
                productos_a_guardar.append(producto_data)

        guardar_productos_en_db(productos_a_guardar)

        return jsonify({
            "datos": productos_a_guardar,
            "num_products": len(productos_a_guardar),
            "processing_time": (datetime.now() - start_time).total_seconds()
        })
    except Exception as e:
        logger.error(f"Error inesperado: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/descargarExcel', methods=['POST', 'OPTIONS'])
def descargar_excel():
    """
    Endpoint para descargar los resultados en formato Excel
    """
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        data = json_util.loads(request.form.get('data'))
        
        # Crear DataFrame con los datos
        df = pd.DataFrame(data['datos'])
        
        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos MercadoLibre"
        
        # Agregar encabezados
        headers = list(df.columns)
        ws.append(headers)
        
        # Agregar datos
        for _, row in df.iterrows():
            ws.append(row.tolist())
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar Excel en memoria
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Enviar archivo
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='productos_mercadolibre.xlsx'
        )
    except Exception as e:
        logger.error(f"Error al generar el archivo Excel: {str(e)}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)